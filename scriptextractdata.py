import os
import pandas as pd
from openpyxl import load_workbook
import os
import zipfile
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook


# Define the folder containing the Excel files
zip_url = "https://environment.data.gov.uk/api/file/download?fileDataSetId=8bc33f70-72a6-4772-ab8e-73d58a06bf6a&fileName=2022_Incineration_Monitoring_Reports.zip"
output_file = "compiled_outputs.xlsx"

print("Downloading zip file...")
response = requests.get(zip_url)
zip_file = BytesIO(response.content)

# Step 2: Extract the zip file
print("Extracting files...")
with zipfile.ZipFile(zip_file, 'r') as zip_ref:
    # Extract all files to a temporary folder
    extract_folder = "extracted_files"
    os.makedirs(extract_folder, exist_ok=True)
    zip_ref.extractall(extract_folder)

data=[]

# Loop through all files in the folder
for file in os.listdir(extract_folder):
    if file.endswith(".xlsm"):  # Check if the file is an Excel file
        file_path = os.path.join(extract_folder, file)
        
        try:
            # Load the workbook and access the specified worksheet
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook["Operational Data"]  # Replace with your sheet name
            
            # Get the values of the specific cells
            capacity_tonnes_pa = sheet["C4"].value
            input_tonnes_pa = sheet["G18"].value
            heat_generated = sheet["G29"].value
            power_generated = sheet["G24"].value
            
            # Append the filename and cell values to the data list
            data.append({
                "File Name": " ".join(file.split()[1:]), 
                "Capacity": capacity_tonnes_pa, 
                "Heat Generated": heat_generated,
                "Power Generated": power_generated,
                "Waste Input": input_tonnes_pa
            })
        
        except Exception as e:
            print(f"Error processing file {file}: {e}")

# Convert the data list into a DataFrame
df = pd.DataFrame(data)

# Save the DataFrame to an Excel file
df.to_excel(output_file, index=False)

print(f"Data has been compiled into {output_file}")